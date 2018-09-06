#!/Users/mabelfan/anaconda3/bin/python3
#爬取厦门人才网一周内的招聘信息

import requests
import bs4
import random
import openpyxl
import json
import os, csv, re
import urllib
import time
import numpy as np

ua_list=[
        "Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.1 (KHTML, like Gecko) Chrome/22.0.1207.1 Safari/537.1",
        "Mozilla/5.0 (X11; CrOS i686 2268.111.0) AppleWebKit/536.11 (KHTML, like Gecko) Chrome/20.0.1132.57 Safari/536.11",
        "Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/536.6 (KHTML, like Gecko) Chrome/20.0.1092.0 Safari/536.6",
        "Mozilla/5.0 (Windows NT 6.2) AppleWebKit/536.6 (KHTML, like Gecko) Chrome/20.0.1090.0 Safari/536.6",
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_13_2) AppleWebKit/604.4.7 (KHTML, like Gecko) Version/11.0.2 Safari/604.4.7",
      ]

def get_headers():
    temp_headers = {
    	'Accept':'application/json, text/javascript, */*; q=0.01',
        'Accept-Encoding':'gzip, deflate, br',
        'Accept-Language':'zh-cn',
        'Connection':'keep-alive',
        'Content-Length':'55',
        'Content-Type':'application/x-www-form-urlencoded; charset=UTF-8',
        'DNT':"1",
        'Cookie':'SEARCH_ID=b5c44beaf06e4d20b7498c6c23d17afc; Hm_lpvt_4233e74dff0ae5bd0a3d81c6ccf756e6=1528702746; Hm_lvt_4233e74dff0ae5bd0a3d81c6ccf756e6=1528610740,1528611956,1528612432,1528681476; LGRID=20180611153902-82c79d65-6d4a-11e8-9446-5254005c3644; LGSID=20180611153902-82c79b68-6d4a-11e8-9446-5254005c3644; PRE_HOST=; PRE_LAND=https%3A%2F%2Fwww.lagou.com%2Fjobs%2Flist_%25E6%2595%25B0%25E6%258D%25AE%25E5%2588%2586%25E6%259E%2590; PRE_SITE=; PRE_UTM=; _ga=GA1.2.778799277.1528610741; _gat=1; _gid=GA1.2.359679147.1528684608; _putrc=6ECE39502780C2F0123F89F2B170EADC; JSESSIONID=ABAAABAAAGGABCB3D146F1B686A172863440CD8E5EB67B1; TG-TRACK-CODE=search_code; index_location_city=%E5%85%A8%E5%9B%BD; gate_login_token=ec8207768c02e8da8c988e9a0fb613346a152fa20d97fe0bf8d4067e91e1cc3e; login=true; unick=%E8%8C%83%E6%A2%85; hasDeliver=0; showExpriedCompanyHome=1; showExpriedIndex=1; showExpriedMyPublish=1; LGUID=20180610140540-4d8a19e1-6c74-11e8-9446-5254005c3644; user_trace_token=20180610140540-4d8a1338-6c74-11e8-9446-5254005c3644',
        'Host':'www.lagou.com',
        'Origin':'https://www.lagou.com',
        'Referer':'https://www.lagou.com/jobs/list_%E6%95%B0%E6%8D%AE%E5%88%86%E6%9E%90?',
        'X-Anit-Forge-Token':'None',
        'X-Anit-Forge-Code':'0',
        'X-Requested-With':'XMLHttpRequest'
    }
    user_agent = random.choice(ua_list)
    temp_headers['User-Agent'] = user_agent
    return temp_headers
# print(get_headers())

def get_requests(url,timeout=5,timeoutRetry=5):
	try:
		res = requests.get(url, get_headers())
		res.raise_for_status()
		print(res)
		soup = bs4.BeautifulSoup(res.text, 'html.parser')
	except Exception as exc:
		print('There was a problem: %s' % exc)
		if timeoutRetry > 0:
			soup = get_requests(url,timeout=5,timeoutRetry=5)
	return soup

# figure out the number of jobs and the number of pages to be searched
def get_num(soup):
	n = 0
	for strs in soup.select('td[align="center"]'):
		while n < 1:
			pat1 = r'共(\d+)条'
			pat2 = r'第1/(\d+)页'
			# 返回的是list中的string，需转换成int
			numJobs = re.compile(pat1).findall(strs.text)
			numPages = re.compile(pat2).findall(strs.text)
			n += 1
	return int(numPages[0]), int(numJobs[0])

def get_job_urls(soup):
	pageCont = soup.select('tr.bg a.a4')
	num = len(pageCont)
	add_urls = []
	for i in range(num):
		add_u =  pageCont[i].attrs['href']
		add_urls.append(add_u)

	add_urls = [url for url in add_urls if url.find('Company') == -1]
	add_urls = list(set(add_urls))
	job_pre_url = 'https://www.xmrc.com.cn'
	job_urls = [job_pre_url + add_url for add_url in add_urls]
	return job_urls,len(job_urls)

def get_table(soupfile):
	tab = soupfile.findAll("table")
	tab_l = tab[3].findAll("table")
	tab_ll = tab_l[5].findAll('table')
	conts = tab_ll[0].get_text()
	try:
		e_cont = tab_ll[4].select(".bmtk")[3].text
	except Exception as e:
		print(e)
	return conts,e_cont

def get_jobTitle(conts):
	pat_jT = r'\xa0招聘职位：(.*)\r\n'
	jobTitle = re.compile(pat_jT).findall(conts)
	if len(jobTitle) > 0:
		return jobTitle[0]
	else:
		return ''

def get_company(conts):
	pat_comp = r'招聘单位：\xa0(.*)\xa0\r\n'
	company = re.compile(pat_comp).findall(conts)
	if len(company) > 0:
		return company[0]
	else:
		return ''

def get_recruit_time(conts):
	pat_time = r'招聘期限：\xa0(.*)\xa0(.*)\r\n'
	time = re.compile(pat_time).findall(conts)
	if len(time) > 0:
		start_day = time[0][0]
		end_day = time[0][1]
		return start_day, end_day
	else:
		return '',''

def get_dept(conts):
	pat_dep = r'招聘部门：\xa0(.*)\r\n'
	dept = re.compile(pat_dep).findall(conts)
	if len(dept) > 0:
		return dept[0]
	else:
		return ''

def get_cta(conts):
	pat_cta = r'联\xa0系\xa0人：\xa0(.*)\r\n'
	cta = re.compile(pat_cta).findall(conts)
	if len(cta) > 0:
		return cta[0]
	else:
		return ''

def get_add(conts):
	pat_add = r'通信地址：\xa0(.*)\r\n'
	add = re.compile(pat_add).findall(conts)
	if len(add) > 0:
		return add[0]
	else:
		return ''

def get_edu(conts):
	pat_edu = r'学历要求：\xa0(.*)\r\n'
	edu_req = re.compile(pat_edu).findall(conts)
	if len(edu_req) > 0:
		return edu_req[0]
	else:
		return ''

def get_sex(conts):
	pat_sex = r'性别要求：\xa0(.*)\r\n'
	sex_req = re.compile(pat_sex).findall(conts)
	if len(sex_req) > 0:
		return sex_req[0]
	else:
		return ''

def get_attr(conts):
	pat_attr = r'职位性质：\xa0(.*)\r\n'
	attr = re.compile(pat_attr).findall(conts)
	if len(attr) > 0:
		return attr[0]
	else:
		return ''

def get_obj(conts):
	pat_obj = r'招聘对象：\xa0(.*)\r\n'
	obj = re.compile(pat_obj).findall(conts)
	if len(obj) > 0:
		return obj[0]
	else:
		return ''

def get_lct(conts):
	pat_lct = r'工作地点：\xa0(.*)\r\n'
	lct = re.compile(pat_lct).findall(conts)
	if len(lct) > 0:
		return lct[0]
	else:
		return ''

def get_salary(conts):
	pat_salary = r'参考月薪：\xa0(.*)\r\n'
	salary = re.compile(pat_salary).findall(conts)
	if len(salary) > 0:
		return salary[0]
	else:
		return ''

def get_wtime(conts):
	pat_wtime = r'上班时间：\r\n\s*(.*)\r\n\s*(\xa0)*(.*)\r\n'
	wtime = re.compile(pat_wtime).findall(conts)
	if len(wtime) > 0:
		time_workday = wtime[0][0]
		day_workweek = wtime[0][1]
		return time_workday, day_workweek
	else:
		return '',''

def get_welfare(conts):
	pat_welfare = r'薪资福利：\xa0(.*)\r\n'
	welfare = re.compile(pat_welfare).findall(conts)
	if len(welfare) > 0:
		return welfare[0]
	else:
		return ''

def get_resp(conts):
	pat_resp = r'岗位职责(.*)\r\n'
	job_resp = re.compile(pat_resp).findall(conts)
	if len(job_resp) > 0:
		return job_resp[0]
	else:
		return ''   

def get_email(e_cont):
	pat_email = r'电子邮件：\xa0(.*)\s*'
	email = re.compile(pat_email).findall(e_cont)
	if len(email) > 0:
		return email[0]
	else:
		return ''

def getinfo(conts,e_cont,timeout=5,timeoutRetry=5):
	job_info = {"职位名称":0,"公司名称":0,"招聘日期":0,"截止日期":0,"招聘部门":0,"联系人":0,\
	"通信地址":0,"学历要求":0,"性别要求":0,"职位性质":0,"招聘对象":0,"工作地点":0,\
	"参考月薪":0,"上班时间":0,"上班天数":0,"薪资福利":0,"岗位职责":0,"电子邮件":0}

	job_info['职位名称'] = get_jobTitle(conts)
	job_info['公司名称'] = get_company(conts)
	job_info['招聘日期'], job_info['截止日期'] = get_recruit_time(conts)
	job_info['招聘部门'] = get_dept(conts)
	job_info['联系人'] = get_cta(conts)
	job_info['通信地址'] = get_add(conts)
	job_info['学历要求'] = get_edu(conts)
	job_info['性别要求'] = get_sex(conts)
	job_info['职位性质'] = get_attr(conts)
	job_info['招聘对象'] = get_obj(conts)
	job_info['工作地点'] = get_lct(conts)
	job_info['参考月薪'] = get_salary(conts)
	job_info['上班时间'], job_info['上班天数'] = get_wtime(conts)
	job_info['薪资福利'] = get_welfare(conts)
	job_info['岗位职责'] = get_resp(conts)
	job_info['电子邮件'] = get_email(e_cont)
	return job_info

# 写入文件中，不同的职位保存在不同的行
def storedata(filename, jobs):
	wb = openpyxl.Workbook()
	sheet = wb.active
	indexes = ['职位名称','公司名称','招聘日期','截止日期','招聘部门','联系人','通信地址','学历要求',\
	'性别要求','职位性质','招聘对象','工作地点','参考月薪','上班时间','上班天数','薪资福利','岗位职责','电子邮件']
	for col in range(1, len(indexes)+1):
		_ = sheet.cell(column = col, row=1,value="{}".format(indexes[col-1]))

	for row in range(2, len(jobs)+2):
		for col in range(1, len(indexes)+1):
			_ = sheet.cell(column=col, row=row, value = "{}".format(jobs[row-2][indexes[col-1]]))
	
	wb.save(filename)


if __name__ == '__main__':
	# crawl the first page to find out the total number of jobs and pages
	url_pre = "https://www.xmrc.com.cn/net/info/resultg.aspx?a=a&g=g&jobtype=&releaseTime=7&searchtype=1&keyword=&sortby=updatetime&ascdesc=Desc&PageIndex="
	first_page = url_pre + str(1)
	soup_1 = get_requests(first_page,timeout=5,timeoutRetry=5)
	numPages,numJobs = get_num(soup_1)
	print('总共有%s页' % numPages,'总共有%s条职位信息' % numJobs)

	# crawl all the pages and jobs
	infos = []
	for i in range(numPages):
		time.sleep(random.random()*5)
		print('开始爬取第%s页'%str(i+1))
		url = url_pre + str(i+1)
		try:
			soup = get_requests(url,timeout=5,timeoutRetry=5)
			job_urls, n = get_job_urls(soup)
			infos_page = []
			for i in range(n):
				try:
					soup_ind = get_requests(job_urls[i],timeout=5,timeoutRetry=5)
					conts,e_cont = get_table(soup_ind)
					job_info = getinfo(conts,e_cont,timeout=5,timeoutRetry=5)
					infos_page.append(job_info)
				except:
					continue
		infos = infos + infos_page
		except:
			continue
	storedata('job_xmrc.xlsx', jobs)	

