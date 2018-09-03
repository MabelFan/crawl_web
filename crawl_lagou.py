#!/Users/mabelfan/anaconda3/bin/python3

import requests
import bs4
import random
import openpyxl
import json
import os, csv
import urllib
import time
from selenium import webdriver



"""
driver = webdriver.Safari()
url = "https://www.lagou.com/jobs/list_?px=new&city=厦门#filterBox"
driver.get(url)
keyword = "数据分析"
search_t = driver.find_element_by_id("keyword")
search_t.send_keys(keyword)

search_b = driver.find_element_by_id("submit")
search_b.click()
"""


ua_list=[
        "Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.1 (KHTML, like Gecko) Chrome/22.0.1207.1 Safari/537.1",
        "Mozilla/5.0 (X11; CrOS i686 2268.111.0) AppleWebKit/536.11 (KHTML, like Gecko) Chrome/20.0.1132.57 Safari/536.11",
        "Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/536.6 (KHTML, like Gecko) Chrome/20.0.1092.0 Safari/536.6",
        "Mozilla/5.0 (Windows NT 6.2) AppleWebKit/536.6 (KHTML, like Gecko) Chrome/20.0.1090.0 Safari/536.6",
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_13_2) AppleWebKit/604.4.7 (KHTML, like Gecko) Version/11.0.2 Safari/604.4.7",
      ]

def get_headers():
    temp_headers = {
    	'Accept':'application/json, text/javascript, */*; q=0.01',
        'Accept-Encoding':'gzip, deflate, br',
        'Accept-Language':'zh-cn',
        'Connection':'keep-alive',
        'Content-Length':'55',
        'Content-Type':'application/x-www-form-urlencoded; charset=UTF-8',
        'DNT':"1",
        'Cookie':'SEARCH_ID=b5c44beaf06e4d20b7498c6c23d17afc; Hm_lpvt_4233e74dff0ae5bd0a3d81c6ccf756e6=1528702746; Hm_lvt_4233e74dff0ae5bd0a3d81c6ccf756e6=1528610740,1528611956,1528612432,1528681476; LGRID=20180611153902-82c79d65-6d4a-11e8-9446-5254005c3644; LGSID=20180611153902-82c79b68-6d4a-11e8-9446-5254005c3644; PRE_HOST=; PRE_LAND=https%3A%2F%2Fwww.lagou.com%2Fjobs%2Flist_%25E6%2595%25B0%25E6%258D%25AE%25E5%2588%2586%25E6%259E%2590; PRE_SITE=; PRE_UTM=; _ga=GA1.2.778799277.1528610741; _gat=1; _gid=GA1.2.359679147.1528684608; _putrc=6ECE39502780C2F0123F89F2B170EADC; JSESSIONID=ABAAABAAAGGABCB3D146F1B686A172863440CD8E5EB67B1; TG-TRACK-CODE=search_code; index_location_city=%E5%85%A8%E5%9B%BD; gate_login_token=ec8207768c02e8da8c988e9a0fb613346a152fa20d97fe0bf8d4067e91e1cc3e; login=true; unick=%E8%8C%83%E6%A2%85; hasDeliver=0; showExpriedCompanyHome=1; showExpriedIndex=1; showExpriedMyPublish=1; LGUID=20180610140540-4d8a19e1-6c74-11e8-9446-5254005c3644; user_trace_token=20180610140540-4d8a1338-6c74-11e8-9446-5254005c3644',
        'Host':'www.lagou.com',
        'Origin':'https://www.lagou.com',
        'Referer':'https://www.lagou.com/jobs/list_%E6%95%B0%E6%8D%AE%E5%88%86%E6%9E%90?',
        'X-Anit-Forge-Token':'None',
        'X-Anit-Forge-Code':'0',
        'X-Requested-With':'XMLHttpRequest'
    }
    user_agent = random.choice(ua_list)
    temp_headers['User-Agent'] = user_agent
    return temp_headers
# print(get_headers())

def post(url, para, headers=None, proxy=None, timeout=5,timeoutRetry=5):
	if not url or not para:
		print("PostError url or para not exit")
		return None
	try:
		response = requests.post(url, data=para, headers=get_headers())
		print(response.status_code)

		#print(response.text)
		if response.status_code == 200 or response.status_code == 302:
			htmlCode = response.text
		else:
			print('2222222222')
			htmlCode = None
	except Exception as e:
		if timeoutRetry > 0:
			htmlCode = post(url=url, para=para, timeoutRetry=(timeoutRetry-1))
			print('333333333')
			htmlCode = None
	return htmlCode


#对获取的json数据进行处理，获取自己需要的信息，获取每个职位数据页数
def getinfo(url, para):
	htmlCode = post(url, para=para, headers=get_headers())  # 获取网页源码
	if htmlCode == None:
		return false
	html_content = json.loads(htmlCode)
	companies = html_content.get('content').get('positionResult').get('result')
	totalCount = html_content.get('content').get('positionResult').get('totalCount')
	pagesize = html_content.get('content').get('pageSize')

	pages = 0
	if int(totalCount)%int(pagesize) == 0:
		pages = int(int(totalCount)/int(pagesize))
	else:
		pages = int(int(totalCount) // int(pagesize)) + 1
	return pages,companies


# 写入文件中，不同的职位保存在不同的行
def storedata(filename, companies):
	wb = openpyxl.Workbook()
	sheet = wb.active
	indexes = ['companyFullName','positionName','positionLables','salary','workYear','createTime','companySize','companyLabelList','city','district']
	for col in range(1, len(indexes)+1):
		_ = sheet.cell(column = col, row=1,value="{}".format(indexes[col-1]))

	for row in range(2, len(companies)+2):
		for col in range(1, len(indexes)+1):
			_ = sheet.cell(column=col, row=row, value = "{}".format(companies[row-2][indexes[col-1]]))
	
	wb.save(filename)

#获取这个职位的信息
if __name__ == '__main__':
	url = 'https://www.lagou.com/jobs/positionAjax.json?'
	para = {'px': 'default','needAddtionalResult': 'false', 'isSchoolJob': 0, 'first': 'true', 'pn': '1',
            'kd':"数据分析"}
	pages,companies = getinfo(url,para)
	# print(pages)   # 上次爬取有206页
	infos = []
	for i in range(pages):
		para['pn'] = str(i+1)
		time.sleep(random.random()*5)
		print('开始爬取第%s页'%str(i+1))
		try:
			pages,companies = getinfo(url,para)
			infos += companies
		except:
			continue
		if companies == None:
			break
	storedata('jobs_logou_t.xlsx', infos)
